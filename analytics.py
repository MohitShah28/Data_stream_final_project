"""
Echo Chambers Analytics Module
Logs community detection results to Excel after every Girvan-Newman run.
"""

import os
import datetime
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

EXCEL_FILE = "echo_chambers_history.xlsx"

HEADERS = [
    "chunk",
    "timestamp",
    "num_communities",
    "modularity_Q",
    "q_strong",
    "community_id",
    "community_size",
    "community_color",
    "community_pct",
    "bridge_user_1",
    "bridge_user_2",
    "bridge_user_3",
    "bridge_user_4",
    "bridge_user_5",
]

# Column widths
COL_WIDTHS = {
    "chunk": 8, "timestamp": 12, "num_communities": 18,
    "modularity_Q": 14, "q_strong": 10, "community_id": 14,
    "community_size": 16, "community_color": 16, "community_pct": 14,
    "bridge_user_1": 14, 
    "bridge_user_2": 14, 
    "bridge_user_3": 14, 
    "bridge_user_4": 14, 
    "bridge_user_5": 14, 
}

def _hex_to_argb(hex_color):
    """Convert #RRGGBB to AARRGGBB for openpyxl."""
    return "FF" + hex_color.lstrip("#").upper()

def _get_or_create_workbook():
    """Load existing workbook or create a new one with headers."""
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Community History"

        # Header row styling
        header_fill = PatternFill("solid", fgColor="2E4057")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        # Column widths
        for col_idx, header in enumerate(HEADERS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 12)

        # Freeze header row
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 35

    return wb, wb.active


def log_snapshot(chunk_idx, communities, Q, bridge_nodes, betweenness, COLORS):
    """
    Log one snapshot to Excel — one row per community.
    Called after every Girvan-Newman run.
    """
    wb, ws = _get_or_create_workbook()

    timestamp = datetime.datetime.now().strftime("%H:%M:%S")
    total_nodes = sum(len(c) for c in communities)
    q_strong = "Yes" if Q >= 0.3 else "No"

    # Prepare bridge user data (top 5)
    bridge_data = []
    for nid in bridge_nodes[:5]:
        bridge_data.append((str(nid), round(betweenness.get(nid, 0), 4)))
    # Pad to 5 if fewer bridge users
    while len(bridge_data) < 5:
        bridge_data.append(("", ""))

    # Thin border for data rows
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # One row per community
    for comm_idx, community in enumerate(communities):
        color_hex = COLORS[comm_idx % len(COLORS)]
        comm_size = len(community)
        comm_pct  = round(comm_size / total_nodes * 100, 1)

        row_data = [
            chunk_idx,
            timestamp,
            len(communities),
            round(Q, 4),
            q_strong,
            comm_idx + 1,
            comm_size,
            color_hex,
            f"{comm_pct}%",
            bridge_data[0][0],
            bridge_data[1][0],
            bridge_data[2][0],
            bridge_data[3][0],
            bridge_data[4][0],
        ]

        row_num = ws.max_row + 1

        # Alternate row background
        bg_color = "F8FAFC" if row_num % 2 == 0 else "FFFFFF"

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = center
            cell.border = border
            cell.font = Font(name="Calibri", size=10)

            # Color the community_color cell with actual community color
            header = HEADERS[col_idx - 1]
            if header == "community_color":
                cell.fill = PatternFill("solid", fgColor=_hex_to_argb(color_hex))
                cell.font = Font(name="Calibri", size=10, color=_hex_to_argb(color_hex))
            elif header == "modularity_Q":
                q_color = "10B981" if Q >= 0.3 else "E76F51"
                cell.font = Font(name="Calibri", size=10, bold=True, color=q_color)
            elif header == "q_strong":
                q_color = "10B981" if Q >= 0.3 else "E76F51"
                cell.font = Font(name="Calibri", size=10, bold=True, color=q_color)
            elif header == "chunk":
                cell.font = Font(name="Calibri", size=10, bold=True, color="2E4057")
            elif header in ("bridge_score_1", "bridge_score_2", "bridge_score_3",
                            "bridge_score_4", "bridge_score_5"):
                cell.font = Font(name="Calibri", size=10, color="9B5DE5")
            else:
                cell.fill = PatternFill("solid", fgColor=bg_color)

    wb.save(EXCEL_FILE)
    print(f"  → Excel log updated: {EXCEL_FILE} ({ws.max_row - 1} rows)")